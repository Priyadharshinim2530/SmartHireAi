# app.py
import json
import os
import click
from functools import wraps
import re
import secrets
from datetime import datetime, timedelta
import io

from markupsafe import Markup
from flask import (
    Flask, render_template, request, redirect, url_for, session, flash, jsonify, abort, send_file
)

from config import Config
from extensions import db, limiter
from models import User, Company, Seeker, Job, Application, PasswordReset
from ats import score_resume, breakdown_to_json, breakdown_from_json

# The openpyxl import is used for generating Excel reports. Import dynamically so
# missing optional dependency doesn't break static analysis or runtime when not used.
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    # Help type checkers/IDE resolve the symbol when openpyxl is available.
    from openpyxl import Workbook  # type: ignore

try:
    import importlib
    _openpyxl = importlib.import_module("openpyxl")
    Workbook = getattr(_openpyxl, "Workbook")
except Exception:
    Workbook = None

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)
limiter.init_app(app)

# Initialize mailer
from mailer import init_app as init_mailer, send_email
init_mailer(app)



def login_required(role=None):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if "user_id" not in session:
                flash("Please log in to continue.", "error")
                return redirect(url_for("login"))
            if role and session.get("role") != role:
                flash("You don't have access to that page.", "error")
                return redirect(url_for("index"))
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return User.query.get(uid)


EMAIL_PATTERN = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")


def is_valid_email(email: str) -> bool:
    return bool(EMAIL_PATTERN.match(email))


def safe_float(value, default=0.0):
    """Convert a value to float, returning *default* on failure."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def validate_password(password: str):
    if len(password) < 10:
        return False, "Passwords must be at least 10 characters long."
    if not re.search(r"[a-z]", password):
        return False, "Passwords must include at least one lowercase letter."
    if not re.search(r"[A-Z]", password):
        return False, "Passwords must include at least one uppercase letter."
    if not re.search(r"[0-9]", password):
        return False, "Passwords must include at least one digit."
    if not re.search(r"[!@#$%^&*()_+\-=[\]{};':\"\\|,.<>/?]", password):
        return False, "Passwords must include at least one special character."
    return True, ""


def get_csrf_token():
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token
    return token


@app.context_processor
def inject_csrf_field():
    def csrf_field():
        return Markup(f"<input type=\"hidden\" name=\"_csrf_token\" value=\"{get_csrf_token()}\">")
    return {"csrf_field": csrf_field}


@app.before_request
def validate_csrf_token():
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        token = request.form.get("_csrf_token") or request.headers.get("X-CSRF-Token")
        stored = session.get("csrf_token") or ""
        if not token or not secrets.compare_digest(token, stored):
            abort(400, "Invalid or missing CSRF token")


def build_resume_text(seeker: Seeker) -> str:
    """Flatten structured resume fields into ATS-friendly plain text."""
    parts = []
    if seeker.full_name:
        parts.append(seeker.full_name)
    if seeker.headline:
        parts.append(seeker.headline)
    if seeker.summary:
        parts.append("SUMMARY\n" + seeker.summary)
    if seeker.skills:
        parts.append("SKILLS\n" + seeker.skills)
    if seeker.work_history:
        try:
            history = json.loads(seeker.work_history)
            lines = ["EXPERIENCE"]
            for job in history:
                lines.append(
                    f"{job.get('title','')} at {job.get('company','')} "
                    f"({job.get('duration','')})\n{job.get('description','')}"
                )
            parts.append("\n".join(lines))
        except (ValueError, TypeError):
            pass
    if seeker.education:
        parts.append("EDUCATION\n" + seeker.education)
    return "\n\n".join(parts)


@app.route("/")
def index():
    open_jobs_count = Job.query.filter_by(status="open").count()
    company_count = Company.query.count()
    seeker_count = Seeker.query.count()
    return render_template(
        "landing.html",
        open_jobs_count=open_jobs_count,
        company_count=company_count,
        seeker_count=seeker_count,
    )


@app.route("/signup", methods=["GET", "POST"])
def signup():
    # Support both the role-toggle form (which POSTS role) and direct links with ?role=...
    if request.method == "POST":
        role = request.form.get("role")
        # Delegate to the existing handlers so validation, email, and redirects remain unchanged
        if role == "company":
            return company_signup()
        return seeker_signup()

    role = request.args.get("role")
    if role == "company":
        return redirect(url_for("company_signup"))
    return redirect(url_for("seeker_signup"))


@app.route("/seeker/signup", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def seeker_signup():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        name = request.form.get("name", "").strip()

        if not email or not password or not name:
            flash("Please fill in all required fields.", "error")
            return redirect(url_for("seeker_signup"))

        if not is_valid_email(email):
            flash("Please enter a valid email address.", "error")
            return redirect(url_for("seeker_signup"))

        if User.query.filter_by(email=email).first():
            flash("An account with that email already exists.", "error")
            return redirect(url_for("seeker_signup"))

        valid, message = validate_password(password)
        if not valid:
            flash(message, "error")
            return redirect(url_for("seeker_signup"))

        user = User(email=email, role="seeker")
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        profile = Seeker(user_id=user.id, full_name=name)
        db.session.add(profile)
        db.session.commit()

        # Send welcome email (best-effort)
        try:
            send_email(
                to_address=email,
                subject="Welcome to SmartHireAI",
                template_base="welcome_email",
                context={"name": name, "email": email},
            )
        except Exception:
            app.logger.exception("Failed to send welcome email to %s", email)

        flash("Account created successfully! Please log in to continue.", "success")
        return redirect(url_for("seeker_login"))

    return render_template("signup.html", portal="seeker")


@app.route("/company/signup", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def company_signup():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        name = request.form.get("name", "").strip()

        if not email or not password or not name:
            flash("Please fill in all required fields.", "error")
            return redirect(url_for("company_signup"))

        if not is_valid_email(email):
            flash("Please enter a valid email address.", "error")
            return redirect(url_for("company_signup"))

        if User.query.filter_by(email=email).first():
            flash("An account with that email already exists.", "error")
            return redirect(url_for("company_signup"))

        valid, message = validate_password(password)
        if not valid:
            flash(message, "error")
            return redirect(url_for("company_signup"))

        user = User(email=email, role="company")
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        profile = Company(user_id=user.id, company_name=name)
        db.session.add(profile)
        db.session.commit()

        # Send welcome email to company account (best-effort)
        try:
            send_email(
                to_address=email,
                subject="Welcome to SmartHireAI (Company)",
                template_base="welcome_email",
                context={"name": name, "email": email},
            )
        except Exception:
            app.logger.exception("Failed to send welcome email to company %s", email)

        flash("Company account created successfully! Please log in to continue.", "success")
        return redirect(url_for("company_login"))

    return render_template("signup.html", portal="company")


@app.route("/login")
def login():
    role = request.args.get("role")
    if role == "company":
        return redirect(url_for("company_login"))
    return redirect(url_for("seeker_login"))


@app.route("/seeker/login", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def seeker_login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = User.query.filter_by(email=email).first()
        if not user or not user.check_password(password) or user.role != "seeker":
            flash("Incorrect email or password. If you do not yet have a Job Seeker account, please sign up before logging in.", "error")
            return redirect(url_for("seeker_login"))
        session["user_id"] = user.id
        session["role"] = user.role
        return redirect(url_for("dashboard_redirect"))

    return render_template("login.html", portal="seeker")


@app.route("/company/login", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def company_login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = User.query.filter_by(email=email).first()
        if not user or not user.check_password(password) or user.role != "company":
            flash("Incorrect email or password. If you do not yet have a Company account, please sign up before logging in.", "error")
            return redirect(url_for("company_login"))
        session["user_id"] = user.id
        session["role"] = user.role
        return redirect(url_for("dashboard_redirect"))

    return render_template("login.html", portal="company")


@app.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if not email or not is_valid_email(email):
            flash("Please enter a valid email address.", "error")
            return redirect(url_for("forgot_password"))

        user = User.query.filter_by(email=email).first()
        if user:
            PasswordReset.query.filter_by(user_id=user.id, used=False).delete()
            token = secrets.token_urlsafe(32)
            reset = PasswordReset(
                user_id=user.id,
                token=token,
                expires_at=datetime.utcnow() + timedelta(hours=1),
            )
            db.session.add(reset)
            db.session.commit()
            reset_url = url_for("reset_password", token=token, _external=True)

            # Determine the display name for the email template
            name = email
            if user.seeker_profile:
                name = user.seeker_profile.full_name
            elif user.company_profile:
                name = user.company_profile.company_name

            try:
                send_email(
                    to_address=email,
                    subject="Reset your SmartHireAI password",
                    template_base="password_reset_email",
                    context={"name": name, "reset_url": reset_url},
                )
            except Exception:
                app.logger.exception("Failed to send password reset email to %s", email)

            app.logger.debug("Password reset requested for %s", email)
        flash(
            "If an account exists for that email, password reset instructions have been sent.",
            "success",
        )
        return redirect(url_for("login"))

    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    reset = PasswordReset.query.filter_by(token=token, used=False).first()
    if not reset or reset.expires_at < datetime.utcnow():
        flash("That password reset link is invalid or has expired.", "error")
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        password = request.form.get("password", "")
        confirm = request.form.get("confirm", "")
        if password != confirm:
            flash("Passwords do not match.", "error")
            return redirect(url_for("reset_password", token=token))
        valid, message = validate_password(password)
        if not valid:
            flash(message, "error")
            return redirect(url_for("reset_password", token=token))

        reset.user.set_password(password)
        reset.used = True
        db.session.commit()
        flash("Your password has been reset. Please log in to continue.", "success")
        return redirect(url_for("login"))

    return render_template("reset_password.html", token=token)


@app.route("/change-password", methods=["GET", "POST"])
@login_required()
def change_password():
    user = current_user()
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not user.check_password(current_password):
            flash("Current password is incorrect.", "error")
            return redirect(url_for("change_password"))
        if new_password != confirm_password:
            flash("New passwords do not match.", "error")
            return redirect(url_for("change_password"))
        valid, message = validate_password(new_password)
        if not valid:
            flash(message, "error")
            return redirect(url_for("change_password"))

        user.set_password(new_password)
        db.session.commit()
        flash("Your password has been changed successfully.", "success")
        return redirect(url_for("dashboard_redirect"))

    return render_template("change_password.html")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    flash("You've been logged out.", "success")
    return redirect(url_for("index"))


@app.route("/dashboard")
@login_required()
def dashboard_redirect():
    if session.get("role") == "company":
        return redirect(url_for("company_dashboard"))
    return redirect(url_for("seeker_dashboard"))


@app.route("/seeker/dashboard")
@login_required("seeker")
def seeker_dashboard():
    seeker = current_user().seeker_profile
    applications = (
        Application.query.filter_by(seeker_id=seeker.id)
        .order_by(Application.applied_at.desc())
        .limit(5)
        .all()
    )
    open_jobs = Job.query.filter_by(status="open").order_by(Job.created_at.desc()).limit(6).all()
    resume_complete = bool(seeker.skills and seeker.summary and seeker.education)
    return render_template(
        "seeker_dashboard.html",
        seeker=seeker,
        applications=applications,
        open_jobs=open_jobs,
        resume_complete=resume_complete,
        total_applications=Application.query.filter_by(seeker_id=seeker.id).count(),
    )


@app.route("/seeker/resume", methods=["GET", "POST"])
@login_required("seeker")
def seeker_resume():
    seeker = current_user().seeker_profile
    if request.method == "POST":
        seeker.full_name = request.form.get("full_name", "").strip()
        seeker.phone = request.form.get("phone", "").strip()
        seeker.headline = request.form.get("headline", "").strip()
        seeker.skills = request.form.get("skills", "").strip()
        seeker.summary = request.form.get("summary", "").strip()
        seeker.education = request.form.get("education", "").strip()
        try:
            seeker.experience_years = float(request.form.get("experience_years", 0) or 0)
        except ValueError:
            seeker.experience_years = 0

        titles = request.form.getlist("job_title[]")
        companies = request.form.getlist("job_company[]")
        durations = request.form.getlist("job_duration[]")
        descriptions = request.form.getlist("job_description[]")
        history = []
        for t, c, d, desc in zip(titles, companies, durations, descriptions):
            if t.strip() or c.strip():
                history.append({"title": t, "company": c, "duration": d, "description": desc})
        seeker.work_history = json.dumps(history)

        seeker.resume_text = build_resume_text(seeker)
        db.session.commit()
        flash("Resume saved.", "success")
        return redirect(url_for("seeker_resume"))

    work_history = []
    if seeker.work_history:
        try:
            work_history = json.loads(seeker.work_history)
        except (ValueError, TypeError):
            work_history = []
    return render_template("seeker_resume_builder.html", seeker=seeker, work_history=work_history)


@app.route("/seeker/jobs")
@login_required("seeker")
def seeker_jobs():
    q = request.args.get("q", "").strip()
    query = Job.query.filter_by(status="open")
    if q:
        like = f"%{q}%"
        query = query.filter(db.or_(Job.title.ilike(like), Job.required_skills.ilike(like)))
    jobs = query.order_by(Job.created_at.desc()).all()

    seeker = current_user().seeker_profile
    applied_job_ids = {
        a.job_id for a in Application.query.filter_by(seeker_id=seeker.id).all()
    }
    return render_template("seeker_jobs.html", jobs=jobs, applied_job_ids=applied_job_ids, q=q)


@app.route("/seeker/jobs/<int:job_id>/apply", methods=["POST"])
@login_required("seeker")
def seeker_apply(job_id):
    job = Job.query.get_or_404(job_id)
    seeker = current_user().seeker_profile

    if not seeker.resume_text:
        flash("Build your resume before applying so recruiters can see your ATS score.", "error")
        return redirect(url_for("seeker_resume"))

    existing = Application.query.filter_by(job_id=job.id, seeker_id=seeker.id).first()
    if existing:
        flash("You've already applied to this job.", "error")
        return redirect(url_for("seeker_jobs"))

    score, breakdown = score_resume(seeker, job)
    application = Application(
        job_id=job.id,
        seeker_id=seeker.id,
        resume_snapshot=seeker.resume_text,
        ats_score=score,
        score_breakdown=breakdown_to_json(breakdown),
    )
    try:
        db.session.add(application)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        # Import here to keep the top-level imports unchanged
        from sqlalchemy.exc import IntegrityError
        if isinstance(exc, IntegrityError):
            flash("You've already applied to this job.", "error")
            return redirect(url_for("seeker_jobs"))
        raise

    # Send application confirmation email (best-effort)
    try:
        send_email(
            to_address=seeker.user.email,
            subject=f"Application received: {job.title}",
            template_base="application_confirmation",
            context={
                "name": seeker.full_name,
                "job_title": job.title,
                "company_name": job.company.company_name,
                "application_id": application.id,
                "applied_at": application.applied_at.strftime("%Y-%m-%d %H:%M UTC"),
            },
        )
    except Exception:
        app.logger.exception("Failed to send application confirmation to %s", seeker.user.email)

    flash(f"Applied! Your ATS match score for this role is {score}/100.", "success")
    return redirect(url_for("seeker_applications"))


@app.route("/seeker/applications")
@login_required("seeker")
def seeker_applications():
    seeker = current_user().seeker_profile
    applications = (
        Application.query.filter_by(seeker_id=seeker.id)
        .order_by(Application.applied_at.desc())
        .all()
    )
    enriched = []
    for a in applications:
        enriched.append({
            "application": a,
            "job": a.job,
            "breakdown": breakdown_from_json(a.score_breakdown),
        })
    return render_template("seeker_applications.html", items=enriched)


@app.route("/company/dashboard")
@login_required("company")
def company_dashboard():
    company = current_user().company_profile
    jobs = Job.query.filter_by(company_id=company.id).order_by(Job.created_at.desc()).all()
    job_ids = [j.id for j in jobs]
    total_applicants = Application.query.filter(Application.job_id.in_(job_ids)).count() if job_ids else 0
    open_jobs = [j for j in jobs if j.status == "open"]
    return render_template(
        "company_dashboard.html",
        company=company,
        jobs=jobs[:5],
        total_jobs=len(jobs),
        open_jobs_count=len(open_jobs),
        total_applicants=total_applicants,
    )


@app.route("/company/jobs")
@login_required("company")
def company_jobs():
    company = current_user().company_profile
    jobs = Job.query.filter_by(company_id=company.id).order_by(Job.created_at.desc()).all()
    counts = {
        j.id: Application.query.filter_by(job_id=j.id).count() for j in jobs
    }
    return render_template("company_manage_jobs.html", jobs=jobs, counts=counts)


@app.route("/company/jobs/new", methods=["GET", "POST"])
@login_required("company")
def company_new_job():
    if request.method == "POST":
        company = current_user().company_profile
        job = Job(
            company_id=company.id,
            title=request.form.get("title", "").strip(),
            description=request.form.get("description", "").strip(),
            required_skills=request.form.get("required_skills", "").strip(),
            location=request.form.get("location", "").strip(),
            job_type=request.form.get("job_type", "Full-time"),
            min_experience=safe_float(request.form.get("min_experience", 0) or 0),
        )
        db.session.add(job)
        db.session.commit()
        flash("Job posted.", "success")
        return redirect(url_for("company_jobs"))
    return render_template("company_post_job.html", job=None)


@app.route("/company/jobs/<int:job_id>/edit", methods=["GET", "POST"])
@login_required("company")
def company_edit_job(job_id):
    company = current_user().company_profile
    job = Job.query.filter_by(id=job_id, company_id=company.id).first_or_404()
    if request.method == "POST":
        job.title = request.form.get("title", "").strip()
        job.description = request.form.get("description", "").strip()
        job.required_skills = request.form.get("required_skills", "").strip()
        job.location = request.form.get("location", "").strip()
        job.job_type = request.form.get("job_type", "Full-time")
        job.min_experience = safe_float(request.form.get("min_experience", 0) or 0)
        db.session.commit()
        flash("Job updated.", "success")
        return redirect(url_for("company_jobs"))
    return render_template("company_post_job.html", job=job)


@app.route("/company/jobs/<int:job_id>/toggle-status", methods=["POST"])
@login_required("company")
def company_toggle_job(job_id):
    company = current_user().company_profile
    job = Job.query.filter_by(id=job_id, company_id=company.id).first_or_404()
    job.status = "closed" if job.status == "open" else "open"
    db.session.commit()
    return redirect(url_for("company_jobs"))


@app.route("/company/jobs/<int:job_id>/delete", methods=["POST"])
@login_required("company")
def company_delete_job(job_id):
    company = current_user().company_profile
    job = Job.query.filter_by(id=job_id, company_id=company.id).first_or_404()
    db.session.delete(job)
    db.session.commit()
    flash("Job deleted.", "success")
    return redirect(url_for("company_jobs"))


@app.route("/company/jobs/<int:job_id>/applicants")
@login_required("company")
def company_applicants(job_id):
    company = current_user().company_profile
    job = Job.query.filter_by(id=job_id, company_id=company.id).first_or_404()
    top_only = request.args.get("top10") == "1"

    applications = Application.query.filter_by(job_id=job.id).order_by(Application.ats_score.desc()).all()
    if top_only:
        applications = applications[:10]

    enriched = []
    for a in applications:
        enriched.append({
            "application": a,
            "seeker": a.seeker,
            "breakdown": breakdown_from_json(a.score_breakdown),
        })
    return render_template("company_applicants.html", job=job, items=enriched, top_only=top_only)


@app.route("/company/applications/<int:app_id>/status", methods=["POST"])
@login_required("company")
def company_update_status(app_id):
    company = current_user().company_profile
    application = Application.query.get_or_404(app_id)
    if application.job.company_id != company.id:
        flash("Not authorized.", "error")
        return redirect(url_for("company_jobs"))
    new_status = request.form.get("status")
    if new_status in ("applied", "shortlisted", "rejected"):
        application.status = new_status
        db.session.commit()

        # Send selection/rejection emails automatically (best-effort)
        try:
            recipient = application.seeker.user.email
            ctx = {
                "name": application.seeker.full_name,
                "job_title": application.job.title,
                "company_name": application.job.company.company_name,
            }
            if new_status == "shortlisted":
                send_email(
                    to_address=recipient,
                    subject=f"Update on your application: {application.job.title}",
                    template_base="selection_email",
                    context=ctx,
                )
            elif new_status == "rejected":
                send_email(
                    to_address=recipient,
                    subject=f"Update on your application: {application.job.title}",
                    template_base="rejection_email",
                    context=ctx,
                )
        except Exception:
            app.logger.exception("Failed to send status email for application %s", application.id)

    return redirect(url_for("company_applicants", job_id=application.job_id))


@app.route("/company/applications/<int:app_id>/invite", methods=["GET", "POST"])
@login_required("company")
def company_invite(app_id):
    company = current_user().company_profile
    application = Application.query.get_or_404(app_id)
    if application.job.company_id != company.id:
        flash("Not authorized.", "error")
        return redirect(url_for("company_jobs"))

    if request.method == "POST":
        date = request.form.get("date", "").strip()
        time = request.form.get("time", "").strip()
        meeting_link = request.form.get("meeting_link", "").strip()
        message = request.form.get("message", "").strip()
        if not date or not time or not meeting_link:
            flash("Please provide date, time, and meeting link.", "error")
            return redirect(url_for("company_invite", app_id=app_id))

        # Send interview invite email (best-effort)
        try:
            send_email(
                to_address=application.seeker.user.email,
                subject=f"Interview invitation: {application.job.title}",
                template_base="interview_invite",
                context={
                    "name": application.seeker.full_name,
                    "job_title": application.job.title,
                    "company_name": company.company_name,
                    "date": date,
                    "time": time,
                    "meeting_link": meeting_link,
                    "message": message,
                },
            )
            flash("Interview invitation sent.", "success")
        except Exception:
            app.logger.exception("Failed to send interview invite for application %s", application.id)
            flash("Failed to send interview invitation.", "error")
        return redirect(url_for("company_applicants", job_id=application.job_id))

    # GET -> render a simple invite form
    return render_template("company_invite_form.html", application=application)


@app.route("/company/analytics")
@login_required("company")
def company_analytics():
    company = current_user().company_profile
    return render_template("company_analytics.html", company=company)


@app.route("/api/company/analytics-data")
@login_required("company")
def api_company_analytics():
    company = current_user().company_profile
    jobs = Job.query.filter_by(company_id=company.id).all()
    job_ids = [j.id for j in jobs]

    applications_per_job = {j.title: Application.query.filter_by(job_id=j.id).count() for j in jobs}

    # High-level counters
    total_jobs = len(jobs)
    open_jobs = len([j for j in jobs if j.status == "open"])
    closed_jobs = len([j for j in jobs if j.status == "closed"])

    status_counts = {"applied": 0, "shortlisted": 0, "rejected": 0}
    score_buckets = {"0-40": 0, "41-60": 0, "61-80": 0, "81-100": 0}
    daily_trend = {}
    monthly_trend = {}
    total_applications = 0

    if job_ids:
        apps = Application.query.filter(Application.job_id.in_(job_ids)).all()
        for a in apps:
            total_applications += 1
            status_counts[a.status] = status_counts.get(a.status, 0) + 1
            # score buckets
            try:
                score = float(a.ats_score or 0)
            except Exception:
                score = 0
            if score <= 40:
                score_buckets["0-40"] += 1
            elif score <= 60:
                score_buckets["41-60"] += 1
            elif score <= 80:
                score_buckets["61-80"] += 1
            else:
                score_buckets["81-100"] += 1
            # daily
            day = a.applied_at.strftime("%Y-%m-%d") if a.applied_at else "unknown"
            daily_trend[day] = daily_trend.get(day, 0) + 1
            # monthly
            month = a.applied_at.strftime("%Y-%m") if a.applied_at else "unknown"
            monthly_trend[month] = monthly_trend.get(month, 0) + 1

    # sort trends by key (date strings)
    daily_trend_sorted = dict(sorted(daily_trend.items()))
    monthly_trend_sorted = dict(sorted(monthly_trend.items()))

    # 'selected' is not a stored status in the current data model. For compatibility include as 0.
    selected_count = 0

    return jsonify({
        "applications_per_job": applications_per_job,
        "status_counts": status_counts,
        "selected_count": selected_count,
        "score_buckets": score_buckets,
        "daily_trend": daily_trend_sorted,
        "monthly_trend": monthly_trend_sorted,
        "total_jobs": total_jobs,
        "total_applications": total_applications,
        "open_jobs": open_jobs,
        "closed_jobs": closed_jobs,
    })


def _ensure_workbook():
    if Workbook is None:
        raise RuntimeError("openpyxl is not available. Please install openpyxl to enable Excel export.")


@app.route("/company/analytics/export")
@login_required("company")
def company_analytics_export():
    """Export analytics as xlsx or csv. Query param: format=xlsx|csv (default xlsx).
    Returns an attachment.
    """
    fmt = (request.args.get("format") or "xlsx").lower()
    company = current_user().company_profile
    # Reuse the API calculation logic by calling the function directly
    data_resp = api_company_analytics()
    data = data_resp.get_json()

    if fmt == "csv":
        # Build a CSV in memory
        out = io.StringIO()
        out.write('Metric,Value\n')
        out.write(f"Total Jobs,{data.get('total_jobs',0)}\n")
        out.write(f"Open Jobs,{data.get('open_jobs',0)}\n")
        out.write(f"Closed Jobs,{data.get('closed_jobs',0)}\n")
        out.write(f"Total Applications,{data.get('total_applications',0)}\n")
        sc = data.get('status_counts',{})
        out.write(f"Applied,{sc.get('applied',0)}\n")
        out.write(f"Shortlisted,{sc.get('shortlisted',0)}\n")
        out.write(f"Rejected,{sc.get('rejected',0)}\n")
        out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode('utf-8')),
                         mimetype='text/csv',
                         download_name=f"{company.company_name}-analytics.csv",
                         as_attachment=True)

    # Default to xlsx
    _ensure_workbook()
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Metric", "Value"])
    ws.append(["Total Jobs", data.get('total_jobs', 0)])
    ws.append(["Open Jobs", data.get('open_jobs', 0)])
    ws.append(["Closed Jobs", data.get('closed_jobs', 0)])
    ws.append(["Total Applications", data.get('total_applications', 0)])
    sc = data.get('status_counts', {})
    ws.append(["Applied", sc.get('applied', 0)])
    ws.append(["Shortlisted", sc.get('shortlisted', 0)])
    ws.append(["Rejected", sc.get('rejected', 0)])

    # Scores sheet
    ws2 = wb.create_sheet(title="Score Distribution")
    ws2.append(["Bucket", "Count"])
    for k, v in (data.get('score_buckets') or {}).items():
        ws2.append([k, v])

    # Monthly trend sheet
    ws3 = wb.create_sheet(title="Monthly Trend")
    ws3.append(["Month", "Applications"])
    for k, v in (data.get('monthly_trend') or {}).items():
        ws3.append([k, v])

    # Applications per job
    ws4 = wb.create_sheet(title="Per Job")
    ws4.append(["Job Title", "Applicants"])
    for k, v in (data.get('applications_per_job') or {}).items():
        ws4.append([k, v])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"{company.company_name}-analytics.xlsx"
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name=filename, as_attachment=True)


@app.route("/company/profile", methods=["GET", "POST"])
@login_required("company")
def company_profile():
    company = current_user().company_profile
    if request.method == "POST":
        company.company_name = request.form.get("company_name", "").strip()
        company.industry = request.form.get("industry", "").strip()
        company.website = request.form.get("website", "").strip()
        company.description = request.form.get("description", "").strip()
        if not company.company_name:
            flash("Company name is required.", "error")
            return redirect(url_for("company_profile"))
        db.session.commit()
        flash("Company profile updated successfully.", "success")
        return redirect(url_for("company_profile"))
    return render_template("company_profile.html", company=company)


@app.route("/company/<int:company_id>")
@login_required()
def company_profile_view(company_id):
    company = Company.query.get_or_404(company_id)
    jobs = Job.query.filter_by(company_id=company.id, status="open").order_by(Job.created_at.desc()).all()
    applied_job_ids = set()
    if session.get("role") == "seeker":
        seeker = current_user().seeker_profile
        applied_job_ids = {
            a.job_id for a in Application.query.filter_by(seeker_id=seeker.id).all()
        }
    return render_template(
        "company_profile_view.html",
        company=company,
        jobs=jobs,
        applied_job_ids=applied_job_ids
    )


@app.cli.command("init-db")
def init_db():
    """Run with: flask --app app init-db"""
    db.create_all()
    print("Database tables created.")


@app.cli.command("test-email")
@click.argument("recipient")
def test_email(recipient):
    """Run with: flask --app app test-email someone@example.com"""
    print("--- SMTP Configuration ---")
    print(f"SMTP_HOST: {app.config.get('SMTP_HOST')}")
    print(f"SMTP_PORT: {app.config.get('SMTP_PORT')}")
    print(f"SMTP_USERNAME: {app.config.get('SMTP_USERNAME')}")
    print(f"SMTP_FROM: {app.config.get('SMTP_FROM')}")
    print(f"SMTP_USE_TLS: {app.config.get('SMTP_USE_TLS')}")
    print(f"SMTP_USE_SSL: {app.config.get('SMTP_USE_SSL')}")
    print("--------------------------")
    print(f"Attempting to send test email to: {recipient}...")

    try:
        success = send_email(
            to_address=recipient,
            subject="SmartHireAI SMTP Test Email",
            template_base="welcome_email",
            context={"name": "Test User", "email": recipient},
        )
        if success:
            print("SUCCESS: Test email sent successfully!")
        else:
            print("FAILURE: send_email() returned False. Check application logs above for detailed SMTP diagnostics.")
    except Exception as e:
        print(f"FAILURE: Exception occurred while sending test email: {e}")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1", port=5000)
